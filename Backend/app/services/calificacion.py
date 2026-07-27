import math
import re
import unicodedata
from datetime import date
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.core.identidad import normalizar_correo, normalizar_documento
from app.models.actividad_evaluativa import ActividadEvaluativa
from app.models.curso import Curso
from app.models.estudiante import Estudiante
from app.models.matricula import Matricula
from app.models.nota import Nota
from app.models.seccion_porcentaje import SeccionPorcentaje
from app.models.usuario import Usuario
from app.repositories.actividad_evaluativa import ActividadEvaluativaRepository
from app.repositories.curso import CursoRepository
from app.repositories.nota import NotaRepository
from app.repositories.seccion_porcentaje import SeccionPorcentajeRepository
from app.services.importacion_excel import ArchivoInvalido, ErrorFila, parsear_notas_xlsx

# --- HU22: plantilla descargable ---

# §7.2: nombre, apellido y correo. SIN documento: un .xlsx descargado sale del
# sistema y circula por correo o WhatsApp, fuera de todo control de acceso, y
# RNF-05 obliga a cuidar los datos personales de menores. El correo institucional
# es dato de contacto que el docente ya maneja, y basta para emparejar de vuelta.
COLUMNAS_PLANTILLA = ("nombre", "apellido", "correo", "calificacion", "comentario")
ANCHOS_PLANTILLA = (18, 18, 34, 14, 30)

# Excel no admite estos caracteres en el nombre de una hoja, y openpyxl no
# siempre avisa: el archivo sale corrupto y no abre.
_CARACTERES_PROHIBIDOS_HOJA = re.compile(r"[\[\]:*?/\\]")
LARGO_MAXIMO_HOJA = 31

MEDIA_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _sanear_nombre_hoja(nombre: str) -> str:
    """Nombre de hoja que Excel acepta: sin caracteres prohibidos y de 31 como máximo."""
    limpio = _CARACTERES_PROHIBIDOS_HOJA.sub(" ", (nombre or "").strip())
    limpio = " ".join(limpio.split())[:LARGO_MAXIMO_HOJA]
    return limpio or "Notas"


def _sanear_trozo_archivo(texto: str) -> str:
    """Trozo de nombre de archivo sin tildes, espacios ni separadores de ruta.

    Un Content-Disposition con caracteres no ASCII se rompe en algunos
    navegadores, así que el nombre se reduce a [A-Za-z0-9_-].
    """
    normalizado = unicodedata.normalize("NFKD", str(texto or "").strip())
    sin_tildes = "".join(c for c in normalizado if not unicodedata.combining(c))
    saneado = re.sub(r"[^A-Za-z0-9]+", "-", sin_tildes).strip("-")
    return saneado or "sin-nombre"


class CalificacionService:

    def __init__(self, session: Session):
        self.session = session
        self.curso_repo = CursoRepository(session)
        self.seccion_repo = SeccionPorcentajeRepository(session)
        self.actividad_repo = ActividadEvaluativaRepository(session)
        self.nota_repo = NotaRepository(session)

    def _validar_pertenencia_curso(self, curso: Curso, usuario: Usuario) -> None:
        # RN-03: un Docente solo puede operar sobre los cursos que dicta él mismo;
        # Administrador no tiene esta restricción.
        if usuario.rol == "Docente" and curso.id_docente != usuario.id_usuario:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tienes permiso sobre este curso")

    def _cursos_visibles(self, usuario: Usuario | None) -> set[int] | None:
        """Ids de curso que el usuario puede leer, o None si no tiene restricción.

        Es la versión de lectura de _validar_pertenencia_curso: los listados no
        reciben un curso concreto que validar, así que necesitan el conjunto
        completo para acotar la consulta. Administrador no tiene restricción;
        Docente ve los cursos que dicta (RN-03) y Estudiante los de su grado y
        año de matrícula (RN-10a).
        """
        if usuario is None or usuario.rol == "Administrador":
            return None
        if usuario.rol == "Docente":
            cursos = self.curso_repo.listar(id_docente=usuario.id_usuario)
        else:
            cursos = self.curso_repo.listar_para_estudiante(usuario.id_usuario)
        return {curso.id_curso for curso in cursos}

    # --- Secciones ---

    def crear_seccion(self, nombre_seccion: str, porcentaje: float, id_curso: int, usuario: Usuario) -> SeccionPorcentaje:
        nombre_limpio = (nombre_seccion or "").strip()
        if not nombre_limpio:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El nombre de la sección no puede estar vacío")

        if porcentaje is None or not math.isfinite(porcentaje) or porcentaje <= 0 or porcentaje > 100:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El porcentaje debe ser un número entre 0 y 100")

        # RN-b: id_curso debe existir
        curso = self.curso_repo.buscar_por_id(id_curso)
        if not curso:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curso no encontrado")

        self._validar_pertenencia_curso(curso, usuario)

        # RN-b (opcional): avisar si la suma de porcentajes del curso supera 100%, sin bloquear
        secciones_existentes = self.seccion_repo.listar(id_curso=id_curso)
        suma_actual = sum(float(seccion.porcentaje) for seccion in secciones_existentes) + float(porcentaje)

        seccion = SeccionPorcentaje(nombre_seccion=nombre_limpio, porcentaje=porcentaje, id_curso=id_curso)
        seccion = self.seccion_repo.crear(seccion)

        if suma_actual > 100:
            seccion.advertencia = f"Las secciones de este curso suman {suma_actual:.2f}%, superan el 100%."

        return seccion

    def listar_secciones(self, id_curso: int | None = None, usuario: Usuario | None = None) -> list[SeccionPorcentaje]:
        return self.seccion_repo.listar(id_curso=id_curso, ids_curso=self._cursos_visibles(usuario))

    # --- Actividades ---

    def crear_actividad(self, nombre: str, fecha: date, id_seccion: int, usuario: Usuario) -> ActividadEvaluativa:
        nombre_limpio = (nombre or "").strip()
        if not nombre_limpio:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El nombre de la actividad no puede estar vacío")

        # RN-c: id_seccion debe existir
        seccion = self.seccion_repo.buscar_por_id(id_seccion)
        if not seccion:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sección no encontrada")

        self._validar_pertenencia_curso(seccion.curso, usuario)

        actividad = ActividadEvaluativa(nombre=nombre_limpio, fecha=fecha, id_seccion=id_seccion)
        return self.actividad_repo.crear(actividad)

    def listar_actividades(self, id_seccion: int | None = None, usuario: Usuario | None = None) -> list[ActividadEvaluativa]:
        return self.actividad_repo.listar(id_seccion=id_seccion, ids_curso=self._cursos_visibles(usuario))

    # --- Notas ---

    def _validar_calificacion(self, calificacion: float) -> None:
        # RN-a: la calificación debe estar entre 0.00 y 5.00 (y no puede ser NaN/Infinity)
        if calificacion is None or not math.isfinite(calificacion) or calificacion < 0 or calificacion > 5:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La calificación debe estar entre 0.00 y 5.00")

    def _validar_estudiante(self, id_estudiante: int) -> None:
        # RN-e: id_estudiante debe existir, tener rol Estudiante y tener fila en Estudiante
        # (Nota.id_estudiante tiene FK contra estudiante, no contra usuario)
        usuario = self.session.get(Usuario, id_estudiante)
        estudiante = self.session.get(Estudiante, id_estudiante)
        if usuario is None or usuario.rol != "Estudiante" or estudiante is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El estudiante debe existir y tener rol Estudiante")

    def _validar_periodo_abierto(self, actividad: ActividadEvaluativa) -> None:
        # RN-d: solo se pueden crear/cargar notas si el período del curso está 'Abierto'
        periodo_estado = actividad.seccion.curso.periodo.estado
        if periodo_estado != "Abierto":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El período académico de este curso no está abierto")

    def _bloquear_nota(self, id_actividad: int, id_estudiante: int) -> None:
        # RN-f: Nota no tiene constraint único en (id_actividad, id_estudiante) en el esquema;
        # se usa un advisory lock transaccional para serializar el upsert entre requests
        # concurrentes sin tener que modificar Database/schemas.sql.
        self.session.execute(
            text("SELECT pg_advisory_xact_lock(:id_actividad, :id_estudiante)"),
            {"id_actividad": id_actividad, "id_estudiante": id_estudiante},
        )

    def _preparar_nota(self, id_actividad: int, id_estudiante: int, calificacion: float, comentario: str | None) -> Nota:
        self._bloquear_nota(id_actividad, id_estudiante)

        nota_existente = self.nota_repo.buscar_por_actividad_y_estudiante(id_actividad, id_estudiante)
        if nota_existente:
            nota_existente.calificacion = calificacion
            nota_existente.comentario = comentario
            self.session.flush()
            return nota_existente

        nota = Nota(
            id_actividad=id_actividad,
            id_estudiante=id_estudiante,
            calificacion=calificacion,
            comentario=comentario,
        )
        return self.nota_repo.agregar(nota)

    def crear_nota(self, id_actividad: int, id_estudiante: int, calificacion: float, comentario: str | None, usuario: Usuario) -> Nota:
        actividad = self.actividad_repo.buscar_por_id(id_actividad)
        if not actividad:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

        self._validar_pertenencia_curso(actividad.seccion.curso, usuario)
        self._validar_calificacion(calificacion)
        self._validar_estudiante(id_estudiante)
        self._validar_periodo_abierto(actividad)

        nota = self._preparar_nota(id_actividad, id_estudiante, calificacion, comentario)
        self.session.commit()
        self.session.refresh(nota)
        return nota

    def cargar_notas_masivo(self, id_actividad: int, notas: list[dict], usuario: Usuario) -> list[Nota]:
        actividad = self.actividad_repo.buscar_por_id(id_actividad)
        if not actividad:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

        self._validar_pertenencia_curso(actividad.seccion.curso, usuario)
        self._validar_periodo_abierto(actividad)

        if not notas:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La lista de notas no puede estar vacía")

        # Fase 1: validar todas las filas antes de escribir nada en la BD.
        # Así, si una fila es inválida, el lote completo falla sin dejar commits parciales.
        for item in notas:
            self._validar_calificacion(item["calificacion"])
            self._validar_estudiante(item["id_estudiante"])

        # Fase 2: preparar todas las notas (add/flush, sin commit) y confirmar el lote entero
        # en una sola transacción.
        resultado = [
            self._preparar_nota(id_actividad, item["id_estudiante"], item["calificacion"], item.get("comentario"))
            for item in notas
        ]

        self.session.commit()
        for nota in resultado:
            self.session.refresh(nota)

        return resultado

    def listar_notas(self, id_actividad: int | None, usuario: Usuario) -> list[Nota]:
        # RN-04: un Estudiante solo puede ver sus propias notas
        id_estudiante_filtro = usuario.id_usuario if usuario.rol == "Estudiante" else None
        # RN-03: un Docente solo puede ver las notas de los cursos que dicta. Sin este
        # filtro, GET /api/notas sin id_actividad devolvía todas las notas del colegio.
        id_docente_filtro = usuario.id_usuario if usuario.rol == "Docente" else None
        return self.nota_repo.listar(
            id_actividad=id_actividad,
            id_estudiante=id_estudiante_filtro,
            id_docente=id_docente_filtro,
        )
   
    def obtener_promedio_estudiante_materia(self, id_estudiante: int, id_materia: int) -> float:
        return self.nota_repo.obtener_promedio_estudiante_materia(id_estudiante, id_materia)
    def obtener_promedio_grupal_materia(self, id_materia: int, usuario: Usuario) -> float:
        # Validamos que el usuario sea un Docente
        if usuario.rol not in ["Docente", "Administrador"]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Solo los docentes pueden ver el promedio grupal de la materia"
            )
        
        # Llamamos al repositorio pasándole el ID de la materia y el ID del profesor
        return self.nota_repo.obtener_promedio_grupal_materia(id_materia, usuario.id_usuario)

    # --- Importación desde Excel (HU22) ---

    def _obtener_actividad_editable(self, id_actividad: int, usuario: Usuario) -> ActividadEvaluativa:
        """Actividad sobre la que el usuario puede escribir notas ahora mismo.

        Es el mismo trío de comprobaciones que hace cargar_notas_masivo, en el
        mismo orden y con los mismos códigos: 404 -> 403 -> 400.
        """
        actividad = self.actividad_repo.buscar_por_id(id_actividad)
        if not actividad:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

        self._validar_pertenencia_curso(actividad.seccion.curso, usuario)
        self._validar_periodo_abierto(actividad)
        return actividad

    def _estudiantes_del_curso(self, curso: Curso) -> list[dict]:
        """RN-k: los matriculados en el grado y año del curso.

        Es exactamente el mismo alcance que ve TablaNotas, así que no se pueden
        importar notas de alguien que no aparece en la tabla. También es la
        frontera de privacidad de todo este flujo: nada de lo que se devuelve al
        docente sale de este conjunto.
        """
        anio = curso.periodo.anio if curso.periodo is not None else date.today().year

        query = (
            select(
                Matricula.id_estudiante,
                Usuario.nombres,
                Usuario.apellidos,
                Usuario.correo,
                Usuario.documento,
            )
            .join(Estudiante, Estudiante.id_estudiante == Matricula.id_estudiante)
            .join(Usuario, Usuario.id_usuario == Estudiante.id_estudiante)
            .where(Matricula.id_grado == curso.id_grado, Matricula.anio == anio)
        )

        return [
            {
                "id_estudiante": fila.id_estudiante,
                "nombre": fila.nombres,
                "apellido": fila.apellidos,
                "correo": fila.correo,
                "documento": fila.documento,
            }
            for fila in self.session.execute(query).all()
        ]

    def _indices_de_identidad(self, estudiantes: list[dict]) -> dict[str, dict]:
        """Un índice por cada clave de RN-j, todos acotados al curso.

        RN-r se aplica también del lado de la base de datos: el documento que se
        guarda desde el registro ya viene normalizado, pero una fila cargada a
        mano puede traer puntos, y comparar "1.023.456.789" con "1023456789"
        falla en silencio.
        """
        indices = {"id_estudiante": {}, "documento": {}, "correo": {}}

        for estudiante in estudiantes:
            indices["id_estudiante"][estudiante["id_estudiante"]] = estudiante

            documento = normalizar_documento(estudiante.get("documento"))
            if documento:
                indices["documento"][documento] = estudiante

            correo = normalizar_correo(estudiante.get("correo"))  # RN-t
            if correo:
                indices["correo"][correo] = estudiante

        return indices

    def _mensaje_sin_emparejar(self, clave: str, valor, anio: int) -> str:
        """Mensaje de una fila cuya identidad no está en el curso.

        A propósito, ninguno confirma si la persona existe fuera del curso: un
        documento de otro grado devuelve lo mismo que uno inexistente. Si no,
        el importador se convierte en un oráculo para enumerar el directorio del
        colegio (§11), que es el problema del hallazgo H3.
        """
        if clave == "documento":
            return (
                f"Ningún estudiante de este curso tiene el documento {valor}. "
                "Si el estudiante ya está registrado pero su ficha no tiene documento, "
                "usa el correo en esa fila o pídele al administrador que la complete."
            )
        if clave == "correo":
            return (
                f"Ningún estudiante de este curso tiene el correo {valor}. "
                "Revisa el correo, o pídele al administrador que cree la cuenta."
            )
        return (
            f"No hay ningún estudiante con id {valor} matriculado en el grado de "
            f"este curso para el año {anio}."
        )

    def previsualizar_importacion_notas(self, id_actividad: int, contenido: bytes, usuario: Usuario) -> dict:
        """Lee un .xlsx y reporta qué se guardaría. No escribe nada (RN-q).

        Es un POST por el tamaño del cuerpo, no por tener efectos: la escritura
        la sigue haciendo cargar_notas_masivo cuando el docente confirma, con su
        única ruta de validación. Así la vía Excel no puede saltarse ninguna
        regla que la vía manual sí aplica.
        """
        # Fallar barato primero: no tiene sentido parsear 1000 filas para un
        # curso ajeno o un periodo cerrado.
        actividad = self._obtener_actividad_editable(id_actividad, usuario)
        curso = actividad.seccion.curso
        anio = curso.periodo.anio if curso.periodo is not None else date.today().year

        try:
            parseo = parsear_notas_xlsx(contenido)
        except ArchivoInvalido as error:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

        estudiantes = self._estudiantes_del_curso(curso)
        indices = self._indices_de_identidad(estudiantes)

        errores: list[ErrorFila] = list(parseo.errores)
        resueltas: list[tuple] = []          # (fila_cruda, estudiante)
        filas_por_estudiante: dict[int, list] = {}

        for fila in parseo.filas:
            estudiante = indices[fila.clave].get(fila.valor)
            if estudiante is None:
                errores.append(ErrorFila(
                    fila=fila.fila,
                    columna=fila.clave,
                    valor=str(fila.valor),
                    mensaje=self._mensaje_sin_emparejar(fila.clave, fila.valor, anio),
                ))
                continue

            resueltas.append((fila, estudiante))
            filas_por_estudiante.setdefault(estudiante["id_estudiante"], []).append(fila)

        # RN-m: dos filas para el mismo estudiante -> las dos son error y no se
        # aplica ninguna. Se comprueba después de resolver la identidad, no
        # sobre el texto del archivo: así también se detecta al estudiante que
        # aparece una vez por documento y otra por correo.
        duplicados = {
            id_estudiante for id_estudiante, filas in filas_por_estudiante.items() if len(filas) > 1
        }

        filas_validas = []
        for fila, estudiante in resueltas:
            if estudiante["id_estudiante"] in duplicados:
                numeros = ", ".join(str(f.fila) for f in filas_por_estudiante[estudiante["id_estudiante"]])
                errores.append(ErrorFila(
                    fila=fila.fila,
                    columna=fila.clave,
                    valor=str(fila.valor),
                    mensaje=(
                        f"{estudiante['nombre']} {estudiante['apellido']} aparece en más de una "
                        f"fila (filas {numeros}). Deja una sola y vuelve a subir el archivo."
                    ),
                ))
                continue

            filas_validas.append({
                "fila": fila.fila,
                "id_estudiante": estudiante["id_estudiante"],
                "calificacion": fila.calificacion,
                "comentario": fila.comentario,
                # Del sistema, no del archivo: el docente confirma contra lo que
                # la base de datos dice que es esa persona.
                "nombre": estudiante["nombre"],
                "apellido": estudiante["apellido"],
            })

        errores.sort(key=lambda error: error.fila)

        return {
            "id_actividad": actividad.id_actividad,
            "actividad": actividad.nombre,
            "total_filas": parseo.total_filas,
            "filas_validas": filas_validas,
            "filas_omitidas": parseo.omitidas,
            "errores": [vars(error) for error in errores],
            "estudiantes_sin_nota": self._estudiantes_sin_nota(
                actividad, estudiantes, mencionados=set(filas_por_estudiante),
            ),
        }

    def _estudiantes_sin_nota(self, actividad: ActividadEvaluativa, estudiantes: list[dict],
                              mencionados: set[int]) -> list[dict]:
        """RN-u: matriculados que el archivo no menciona y que siguen sin nota.

        Es el único caso que ningún error de fila puede detectar: el estudiante
        que el archivo *no* nombra. Un docente que borró una fila sin darse
        cuenta vería "10 notas guardadas" y todo parecería correcto. Cuenta
        también las notas que ya estaban cargadas antes: si a Sara se la
        pusieron a mano, no aparece.

        Es un aviso, no un error: importar media clase es legítimo.
        """
        con_nota_previa = {
            nota.id_estudiante for nota in self.nota_repo.listar(id_actividad=actividad.id_actividad)
        }

        return [
            {
                "id_estudiante": estudiante["id_estudiante"],
                "nombre": estudiante["nombre"],
                "apellido": estudiante["apellido"],
            }
            for estudiante in estudiantes
            if estudiante["id_estudiante"] not in mencionados
            and estudiante["id_estudiante"] not in con_nota_previa
        ]

    def generar_plantilla_notas(self, id_actividad: int, usuario: Usuario) -> tuple[bytes, str]:
        """Un .xlsx con la lista real del curso y la columna de nota vacía.

        Es lo que hace que el emparejamiento deje de ser un problema en el flujo
        principal: el docente no tiene que escribir ninguna clave, solo las
        notas. Devuelve (contenido, nombre de archivo).

        El periodo cerrado no bloquea la descarga: consultar no es escribir. El
        bloqueo aplica al importar y al confirmar (RN-d).
        """
        actividad = self.actividad_repo.buscar_por_id(id_actividad)
        if not actividad:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

        curso = actividad.seccion.curso
        self._validar_pertenencia_curso(curso, usuario)

        estudiantes = self._estudiantes_del_curso(curso)
        notas_actuales = {
            nota.id_estudiante: float(nota.calificacion)
            for nota in self.nota_repo.listar(id_actividad=id_actividad)
        }

        libro = Workbook()
        hoja = libro.active
        # El nombre de la hoja dice a qué actividad pertenece el archivo sin
        # meter el id_actividad en los datos (RN-w). Es ayuda visual: al
        # importar no se lee, la verificación real la hace el frontend (§10.2).
        hoja.title = _sanear_nombre_hoja(actividad.nombre)

        hoja.append(list(COLUMNAS_PLANTILLA))
        for indice, ancho in enumerate(ANCHOS_PLANTILLA, start=1):
            hoja.cell(row=1, column=indice).font = Font(bold=True)
            hoja.column_dimensions[get_column_letter(indice)].width = ancho

        # Con 40 estudiantes el docente pierde de vista qué columna está llenando.
        hoja.freeze_panes = "A2"

        estudiantes.sort(key=lambda e: ((e["apellido"] or "").lower(), (e["nombre"] or "").lower()))
        for estudiante in estudiantes:
            hoja.append([
                estudiante["nombre"],
                estudiante["apellido"],
                estudiante["correo"],
                notas_actuales.get(estudiante["id_estudiante"]),
                None,
            ])
            # El correo como texto: evita que Excel lo convierta en hipervínculo
            # y que algún copiar/pegar arrastre el enlace en vez del texto.
            hoja.cell(row=hoja.max_row, column=3).number_format = "@"

        buffer = BytesIO()
        libro.save(buffer)
        libro.close()

        nombre_archivo = "notas-{}-{}-{}.xlsx".format(
            _sanear_trozo_archivo(curso.materia.nombre if curso.materia else "materia"),
            _sanear_trozo_archivo(curso.grado.nombre if curso.grado else "grado"),
            _sanear_trozo_archivo(actividad.nombre),
        )

        return buffer.getvalue(), nombre_archivo

    # --- Eliminaciones (HU16) ---
    def eliminar_actividad(self, id_actividad: int, usuario: Usuario) -> None:
        actividad = self.actividad_repo.buscar_por_id(id_actividad)
        if not actividad:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

        # Validar pertenencia y periodo abierto
        self._validar_pertenencia_curso(actividad.seccion.curso, usuario)
        self._validar_periodo_abierto(actividad)

        # Borrar notas asociadas y la actividad
        self.nota_repo.borrar_por_actividad(actividad.id_actividad)
        self.actividad_repo.borrar(actividad)
        self.session.commit()

    def eliminar_seccion(self, id_seccion: int, usuario: Usuario) -> None:
        seccion = self.seccion_repo.buscar_por_id(id_seccion)
        if not seccion:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sección no encontrada")

        # RN-03: validar pertenencia de curso
        self._validar_pertenencia_curso(seccion.curso, usuario)

        # Solo permitir en período abierto
        periodo_estado = seccion.curso.periodo.estado
        if periodo_estado != "Abierto":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El período académico de este curso no está abierto")

        # Borrar actividades y sus notas
        actividades = self.actividad_repo.listar(id_seccion=seccion.id_seccion)
        for actividad in actividades:
            self.nota_repo.borrar_por_actividad(actividad.id_actividad)
            self.actividad_repo.borrar(actividad)

        # Borrar la sección
        self.seccion_repo.borrar(seccion)
        self.session.commit()