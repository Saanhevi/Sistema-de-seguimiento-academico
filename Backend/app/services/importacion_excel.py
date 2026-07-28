"""Traduce un .xlsx de notas a filas crudas. No sabe qué es una nota.

Este módulo es deliberadamente puro: sin SQLAlchemy, sin FastAPI y sin ninguna
noción de calificaciones más allá de que hay una celda con un número entre 0 y 5.
Recibe bytes y devuelve estructuras de datos. Eso permite dos cosas:

  - probarlo con archivos generados en memoria, sin Postgres ni cliente HTTP;
  - reutilizarlo tal cual si alguna vez se importan estudiantes desde Excel
    (HU20), sin arrastrar el módulo de notas detrás.

Quien resuelve a qué estudiante corresponde cada fila es CalificacionService:
aquí solo se dice "esta fila trae el documento 1023456789", nunca "esta fila es
de Ana Gómez".

Reglas implementadas: RN-g (firma ZIP), RN-h (límites), RN-i (encabezados),
RN-j (precedencia de claves), RN-l (celda vacía), RN-n (separador decimal),
RN-o (redondeo), RN-p (comentario), RN-r (normalización), RN-s (notación
científica), RN-t (correo).
"""

import math
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

from app.core.identidad import es_notacion_cientifica, normalizar_correo, normalizar_documento

# RN-h. El tamaño se comprueba sobre los bytes recibidos, antes de abrir nada:
# un .xlsx es un ZIP y un archivo pequeño puede descomprimirse a gigas.
TAMANO_MAXIMO_BYTES = 2 * 1024 * 1024
FILAS_MAXIMAS = 1000

# RN-g. Se valida por contenido y no por la extensión del nombre ni por el
# Content-Type, que los pone el cliente y puede mentir en los dos.
FIRMA_ZIP = b"PK\x03\x04"

# Encabezados aceptados -> nombre canónico de la columna (§7.1).
ALIAS_COLUMNAS = {
    "id_estudiante": "id_estudiante",
    "id": "id_estudiante",
    "documento": "documento",
    "identificacion": "documento",
    "cedula": "documento",
    "num_documento": "documento",
    "no_documento": "documento",
    "correo": "correo",
    "email": "correo",
    "correo_electronico": "correo",
    "calificacion": "calificacion",
    "nota": "calificacion",
    "comentario": "comentario",
    "observacion": "comentario",
    "nombre": "nombre",
    "apellido": "apellido",
}

# RN-j: precedencia de emparejamiento. Se usa la primera clave presente y no
# vacía en la fila; si esa falla, la fila es error y NO se reintenta con la
# siguiente. Probar todas las claves hasta que alguna pegue es cómo se le
# termina poniendo la nota de un estudiante a otro.
CLAVES_IDENTIDAD = ("id_estudiante", "documento", "correo")

COMENTARIO_MAXIMO = 100  # Nota.comentario es VARCHAR(100)

# La primera fila de datos es la 2 en Excel: la 1 son los encabezados. Los
# mensajes de error usan este número y no un índice base 0, porque un mensaje
# que dice "fila 5" cuando en Excel es la 6 es peor que no decir nada.
PRIMERA_FILA_DATOS = 2


class ArchivoInvalido(Exception):
    """El archivo completo no se puede procesar; no hay filas que reportar.

    Se distingue de un ErrorFila a propósito: esto termina en un 400 y el
    docente tiene que arreglar el archivo entero, no una celda.
    """


@dataclass(frozen=True)
class ErrorFila:
    """Un problema concreto, ubicable en el archivo del docente."""

    fila: int
    columna: str
    valor: str | None
    mensaje: str


@dataclass(frozen=True)
class FilaCruda:
    """Una fila con nota, ya normalizada, todavía sin resolver a un estudiante."""

    fila: int
    clave: str          # cuál de CLAVES_IDENTIDAD se usó (RN-j)
    valor: str | int    # el valor de esa clave, ya normalizado
    calificacion: float
    comentario: str | None = None
    nombre: str | None = None
    apellido: str | None = None


@dataclass
class ResultadoParseo:
    """Lo que el parser puede decir de un archivo sin consultar la base de datos.

    `omitidas` no cabe en filas ni en errores (RN-l: una celda de calificación
    vacía no es ninguna de las dos cosas) y el frontend la necesita para el
    resumen, así que el resultado es un objeto y no la tupla de dos listas.
    """

    filas: list[FilaCruda] = field(default_factory=list)
    errores: list[ErrorFila] = field(default_factory=list)
    omitidas: int = 0

    @property
    def total_filas(self) -> int:
        return len(self.filas) + len(self.errores) + self.omitidas


def _normalizar_encabezado(valor) -> str:
    """'Nº Documento ' -> 'n_documento'. Sin tildes, en minúsculas, sin espacios."""
    if valor is None:
        return ""
    texto = unicodedata.normalize("NFKD", str(valor).strip().lower())
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    # Cualquier cosa que no sea letra o dígito se vuelve separador; los
    # separadores seguidos colapsan en uno solo y no quedan en los extremos.
    partes = ["".join(c for c in trozo if c.isalnum()) for trozo in texto.replace("_", " ").split()]
    return "_".join(parte for parte in partes if parte)


def _texto(valor) -> str | None:
    """Valor de celda a texto limpio, o None si la celda está vacía."""
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        # Excel entrega las celdas numéricas como float. Sin este paso, el
        # documento 1023456789 llegaría como "1023456789.0" y no emparejaría
        # nunca (RN-r).
        valor = int(valor)
    texto = str(valor).strip()
    return texto or None


def _mapear_encabezados(fila_encabezados) -> dict[str, int]:
    """{columna canónica: índice de celda}. Las columnas desconocidas se ignoran."""
    mapa: dict[str, int] = {}
    for indice, celda in enumerate(fila_encabezados or ()):
        canonico = ALIAS_COLUMNAS.get(_normalizar_encabezado(celda))
        # Primera aparición gana: si el docente dejó dos columnas "nota", la de
        # más a la izquierda es la que estaba viendo mientras escribía.
        if canonico and canonico not in mapa:
            mapa[canonico] = indice
    return mapa


def _validar_encabezados(mapa: dict[str, int]) -> None:
    # RN-i: sin una columna de identidad no hay forma de saber de quién es cada
    # nota. No se empareja por nombre (§2.2), así que el archivo se rechaza
    # entero en vez de reportar mil filas con el mismo error.
    if not any(clave in mapa for clave in CLAVES_IDENTIDAD):
        raise ArchivoInvalido(
            "El archivo debe tener al menos una columna de identidad: "
            "id_estudiante, documento o correo."
        )
    if "calificacion" not in mapa:
        raise ArchivoInvalido(
            "El archivo debe tener una columna de calificación (encabezado "
            "'calificacion' o 'nota')."
        )


def _celda(fila, mapa: dict[str, int], columna: str):
    indice = mapa.get(columna)
    if indice is None or indice >= len(fila):
        return None
    return fila[indice]


def _leer_identidad(fila, mapa, numero_fila) -> tuple[str, str | int] | ErrorFila:
    """Aplica RN-j: la primera clave presente manda, y si falla no hay segunda oportunidad."""
    for clave in CLAVES_IDENTIDAD:
        bruto = _celda(fila, mapa, clave)
        texto = _texto(bruto)
        if texto is None:
            continue

        if clave == "id_estudiante":
            try:
                valor = int(float(texto))
            except (TypeError, ValueError):
                return ErrorFila(numero_fila, "id_estudiante", texto,
                                 "El id del estudiante debe ser un número entero.")
            if valor <= 0:
                return ErrorFila(numero_fila, "id_estudiante", texto,
                                 "El id del estudiante debe ser un número entero positivo.")
            return clave, valor

        if clave == "documento":
            # RN-s: "1,02E+09" no se adivina. Excel ya perdió los dígitos que
            # faltan; reconstruir el número sería inventárselo.
            if es_notacion_cientifica(texto):
                return ErrorFila(numero_fila, "documento", texto,
                                 "El documento llegó en notación científica. Formatea la columna "
                                 "documento como texto en Excel y vuelve a exportar el archivo.")
            if isinstance(bruto, float) and not bruto.is_integer():
                return ErrorFila(numero_fila, "documento", texto,
                                 "El documento no parece un número de identificación. Formatea la "
                                 "columna documento como texto en Excel y vuelve a exportar.")
            normalizado = normalizar_documento(texto)
            if not normalizado:
                continue
            return clave, normalizado

        normalizado = normalizar_correo(texto)  # RN-t
        if not normalizado:
            continue
        return clave, normalizado

    return ErrorFila(
        numero_fila, "identidad", None,
        "La fila no trae id_estudiante, documento ni correo, así que no hay forma de "
        "saber de qué estudiante es la nota.",
    )


def _leer_calificacion(bruto, numero_fila) -> float | ErrorFila | None:
    """None = celda vacía (RN-l). El resto: float redondeado a 2 decimales o error."""
    if bruto is None or (isinstance(bruto, str) and not bruto.strip()):
        return None

    if isinstance(bruto, bool):
        # bool es subclase de int en Python; sin esto, TRUE valdría 1.00.
        return ErrorFila(numero_fila, "calificacion", str(bruto),
                         "La calificación debe ser un número entre 0.00 y 5.00.")

    if isinstance(bruto, (int, float)):
        valor = float(bruto)
    else:
        # RN-n: solo las celdas de texto necesitan esto; las numéricas ya llegan
        # como float. El docente colombiano escribe "4,5" tanto como "4.5".
        try:
            valor = float(str(bruto).strip().replace(",", "."))
        except (TypeError, ValueError):
            return ErrorFila(numero_fila, "calificacion", str(bruto).strip(),
                             "La calificación debe ser un número entre 0.00 y 5.00.")

    if not math.isfinite(valor) or valor < 0 or valor > 5:
        return ErrorFila(numero_fila, "calificacion", str(bruto).strip(),
                         "La calificación debe estar entre 0.00 y 5.00.")

    # RN-o: la columna es NUMERIC(3,2). Se redondea aquí para que el valor que
    # el docente aprueba en la vista previa sea exactamente el que se guarda,
    # en vez de dejar que Postgres lo redondee de forma invisible.
    return round(valor, 2)


def _leer_comentario(bruto, numero_fila) -> str | ErrorFila | None:
    texto = _texto(bruto)
    if texto is None:
        return None
    if len(texto) > COMENTARIO_MAXIMO:
        # RN-p: truncar en silencio altera lo que el docente escribió.
        return ErrorFila(numero_fila, "comentario", texto[:40] + "…",
                         f"El comentario supera los {COMENTARIO_MAXIMO} caracteres "
                         f"(tiene {len(texto)}). Acórtalo y vuelve a subir el archivo.")
    return texto


def _validar_archivo(contenido: bytes) -> None:
    if not contenido:
        raise ArchivoInvalido("El archivo está vacío.")
    # RN-h antes que RN-g: comprobar el tamaño no requiere interpretar nada.
    if len(contenido) > TAMANO_MAXIMO_BYTES:
        raise ArchivoInvalido(
            f"El archivo pesa más de {TAMANO_MAXIMO_BYTES // (1024 * 1024)} MB. "
            "Sube solo las notas de una actividad."
        )
    if not contenido.startswith(FIRMA_ZIP):
        raise ArchivoInvalido(
            "El archivo no es un Excel .xlsx válido. Si tu archivo es .xls o .csv, "
            "ábrelo en Excel y guárdalo como .xlsx."
        )


def parsear_notas_xlsx(contenido: bytes) -> ResultadoParseo:
    """Traduce un .xlsx a filas crudas. No sabe qué es una nota ni consulta la BD.

    Lanza ArchivoInvalido si el archivo entero no sirve (RN-g/RN-h/RN-i); los
    problemas de una fila concreta nunca lanzan, se acumulan en `errores`.
    Solo se lee la primera hoja del libro.
    """
    _validar_archivo(contenido)

    workbook = None
    try:
        # read_only: no carga la hoja entera en memoria.
        # data_only: lee el *valor* de las fórmulas y no la fórmula en texto.
        workbook = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        hoja = workbook.worksheets[0]
        filas = hoja.iter_rows(values_only=True)

        try:
            encabezados = next(filas)
        except StopIteration:
            raise ArchivoInvalido("El archivo no tiene ninguna fila; se esperaba una de encabezados.")

        mapa = _mapear_encabezados(encabezados)
        _validar_encabezados(mapa)

        resultado = ResultadoParseo()
        filas_con_datos = 0

        for desplazamiento, fila in enumerate(filas):
            numero_fila = PRIMERA_FILA_DATOS + desplazamiento

            # Fila totalmente vacía: Excel las arrastra al final del rango sin
            # que nadie las escriba. Ni cuentan para el límite ni son omitidas.
            if fila is None or all(_texto(celda) is None for celda in fila):
                continue

            filas_con_datos += 1
            if filas_con_datos > FILAS_MAXIMAS:
                raise ArchivoInvalido(
                    f"El archivo tiene más de {FILAS_MAXIMAS} filas de datos. "
                    "Sube las notas de un curso a la vez."
                )

            calificacion = _leer_calificacion(_celda(fila, mapa, "calificacion"), numero_fila)
            if isinstance(calificacion, ErrorFila):
                resultado.errores.append(calificacion)
                continue
            if calificacion is None:
                # RN-l: sin nota no hay nada que guardar y no hay nada que
                # reportar. Permite subir la plantilla a medio llenar. No se
                # valida siquiera la identidad: la fila no va a producir nada.
                resultado.omitidas += 1
                continue

            identidad = _leer_identidad(fila, mapa, numero_fila)
            if isinstance(identidad, ErrorFila):
                resultado.errores.append(identidad)
                continue

            comentario = _leer_comentario(_celda(fila, mapa, "comentario"), numero_fila)
            if isinstance(comentario, ErrorFila):
                resultado.errores.append(comentario)
                continue

            clave, valor = identidad
            resultado.filas.append(FilaCruda(
                fila=numero_fila,
                clave=clave,
                valor=valor,
                calificacion=calificacion,
                comentario=comentario,
                # Informativas: hacen legible la vista previa, no emparejan (§2.2).
                nombre=_texto(_celda(fila, mapa, "nombre")),
                apellido=_texto(_celda(fila, mapa, "apellido")),
            ))

        return resultado

    except ArchivoInvalido:
        raise
    except Exception as error:
        # Un .xlsx corrupto hace que openpyxl lance excepciones de todos los
        # colores (BadZipFile, KeyError, TypeError...). Se captura amplio para
        # que cualquiera de ellas salga como un 400 entendible y nunca como un
        # 500 con un stack trace.
        raise ArchivoInvalido(
            "No se pudo leer el archivo: parece dañado o no es un Excel válido."
        ) from error
    finally:
        if workbook is not None:
            # En modo read_only openpyxl deja descriptores abiertos.
            workbook.close()
