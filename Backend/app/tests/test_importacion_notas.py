"""
Pruebas de la importación de notas en CalificacionService (HU22).

El parser ya está cubierto en test_importacion_excel.py; aquí se prueba lo que
solo el servicio sabe hacer: resolver cada fila a un estudiante del curso y
decidir qué se puede guardar.

  - Orden de validación: 404 -> 403 -> 400 periodo, antes de leer el archivo.
  - RN-k: solo estudiantes matriculados en el grado y año del curso.
  - RN-j/RN-r: emparejamiento por documento, normalizado por los dos lados.
  - RN-m: dos filas del mismo estudiante -> las dos son error.
  - RN-q: la previsualización no escribe nada.
  - RN-u: aviso de matriculados que el archivo no menciona.
  - §11: no confirma la existencia de nadie fuera del curso.
  - §7.2: la plantilla lleva correo y NO lleva documento.

Todo con dobles de prueba; no se toca la base de datos.
"""

import os
import unittest
from io import BytesIO
from unittest.mock import Mock

os.environ.setdefault("DATABASE_URL", "postgresql+psycopg://postgres:postgres@localhost:5433/gestion_academica")
os.environ.setdefault("SECRET_KEY", "dev_secret_key_local_only")
os.environ.setdefault("ALGORITHM", "HS256")
os.environ.setdefault("ACCESS_TOKEN_EXPIRE_MINUTES", "60")

from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.dependencies import get_calificacion_service, get_current_user
from app.main import app
from app.models.usuario import Usuario
from app.services.calificacion import MEDIA_TYPE_XLSX, CalificacionService
from app.tests.test_importacion_excel import construir_xlsx

DOCENTE = Usuario(id_usuario=3, rol="Docente")
AJENO = Usuario(id_usuario=99, rol="Docente")

# El curso: 6°A de Matemáticas, año 2026. Tres estudiantes matriculados.
# Sara no tiene documento cargado: es el caso de los usuarios creados antes de
# que existiera la columna (§3.4).
ESTUDIANTES = [
    {"id_estudiante": 11, "nombre": "Ana", "apellido": "Gómez",
     "correo": "ana.gomez@colegio.edu.co", "documento": "1.023.456.789"},
    {"id_estudiante": 12, "nombre": "Luis", "apellido": "Peña",
     "correo": "Luis.Pena@Colegio.edu.co", "documento": "0012345678"},
    {"id_estudiante": 14, "nombre": "Sara", "apellido": "Ríos",
     "correo": "sara.rios@colegio.edu.co", "documento": None},
]


def _actividad(id_docente=3, estado_periodo="Abierto", nombre="Parcial 1"):
    actividad = Mock()
    actividad.id_actividad = 7
    actividad.nombre = nombre

    curso = Mock()
    curso.id_curso = 5
    curso.id_docente = id_docente
    curso.id_grado = 2
    curso.periodo = Mock(estado=estado_periodo, anio=2026)
    curso.materia = Mock()
    curso.materia.nombre = "Matemáticas"
    curso.grado = Mock()
    curso.grado.nombre = "3°A"

    actividad.seccion = Mock()
    actividad.seccion.curso = curso
    return actividad


def _servicio(actividad=None, estudiantes=None, notas_existentes=()):
    """CalificacionService con los repositorios y la consulta del curso simulados."""
    service = CalificacionService(Mock())
    service.actividad_repo = Mock()
    service.actividad_repo.buscar_por_id.return_value = actividad
    service.nota_repo = Mock()
    service.nota_repo.listar.return_value = [
        Mock(id_estudiante=id_estudiante, calificacion=calificacion)
        for id_estudiante, calificacion in notas_existentes
    ]
    # La consulta del listado es de CursoService y ya está probada allí; aquí
    # interesa qué hace el servicio con el resultado.
    service._estudiantes_del_curso = Mock(
        return_value=[dict(e) for e in (ESTUDIANTES if estudiantes is None else estudiantes)]
    )
    return service


def _archivo(filas):
    return construir_xlsx(filas)


class OrdenDeValidacionTests(unittest.TestCase):
    """§8.2: fallar barato primero."""

    def test_actividad_inexistente_devuelve_404(self):
        service = _servicio(actividad=None)
        with self.assertRaises(HTTPException) as exc:
            service.previsualizar_importacion_notas(7, b"lo que sea", DOCENTE)
        self.assertEqual(exc.exception.status_code, 404)

    def test_docente_ajeno_devuelve_403(self):
        service = _servicio(_actividad(id_docente=3))
        with self.assertRaises(HTTPException) as exc:
            service.previsualizar_importacion_notas(7, b"lo que sea", AJENO)
        self.assertEqual(exc.exception.status_code, 403)

    def test_periodo_cerrado_devuelve_400_sin_mirar_el_archivo(self):
        service = _servicio(_actividad(estado_periodo="Cerrado"))
        # Bytes que no son un .xlsx: si el archivo se llegara a parsear, el
        # mensaje sería el del parser y no el del periodo.
        with self.assertRaises(HTTPException) as exc:
            service.previsualizar_importacion_notas(7, b"no soy un excel", DOCENTE)
        self.assertEqual(exc.exception.status_code, 400)
        self.assertIn("período", exc.exception.detail)

    def test_archivo_invalido_es_400_y_no_500(self):
        service = _servicio(_actividad())
        with self.assertRaises(HTTPException) as exc:
            service.previsualizar_importacion_notas(7, b"no soy un excel", DOCENTE)
        self.assertEqual(exc.exception.status_code, 400)

    def test_administrador_no_tiene_restriccion_de_curso(self):
        service = _servicio(_actividad(id_docente=3))
        admin = Usuario(id_usuario=1, rol="Administrador")
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["ana.gomez@colegio.edu.co", 4.5]]), admin
        )
        self.assertEqual(len(resultado["filas_validas"]), 1)


class EmparejamientoTests(unittest.TestCase):
    """RN-j, RN-k, RN-r, RN-t."""

    def test_empareja_por_correo_sin_importar_mayusculas(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["LUIS.PENA@colegio.edu.co", "3,2"]]), DOCENTE
        )
        self.assertEqual(resultado["errores"], [])
        fila = resultado["filas_validas"][0]
        self.assertEqual(fila["id_estudiante"], 12)
        self.assertEqual(fila["calificacion"], 3.2)

    def test_empareja_por_documento_normalizando_los_dos_lados(self):
        """RN-r: en la BD está "1.023.456.789" y el archivo trae "1023456789"."""
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["documento", "calificacion"], ["1023456789", 4.5]]), DOCENTE
        )
        self.assertEqual(resultado["errores"], [])
        self.assertEqual(resultado["filas_validas"][0]["id_estudiante"], 11)

    def test_documento_con_ceros_a_la_izquierda_no_pierde_digitos(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["documento", "calificacion"], ["0012345678", 4]]), DOCENTE
        )
        self.assertEqual(resultado["filas_validas"][0]["id_estudiante"], 12)

    def test_nombre_y_apellido_salen_del_sistema_no_del_archivo(self):
        # El docente confirma contra lo que la base de datos dice que es esa
        # persona, no contra lo que él escribió en su planilla.
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["nombre", "apellido", "correo", "calificacion"],
                      ["Anna", "Gomes", "ana.gomez@colegio.edu.co", 4]]),
            DOCENTE,
        )
        fila = resultado["filas_validas"][0]
        self.assertEqual((fila["nombre"], fila["apellido"]), ("Ana", "Gómez"))

    def test_estudiante_fuera_del_curso_es_error_de_fila(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["ajeno@colegio.edu.co", 4]]), DOCENTE
        )
        self.assertEqual(resultado["filas_validas"], [])
        self.assertEqual(resultado["errores"][0]["columna"], "correo")
        self.assertEqual(resultado["errores"][0]["fila"], 2)

    def test_el_mensaje_no_confirma_que_la_persona_exista_fuera_del_curso(self):
        """§11: si no, el importador sirve para enumerar el directorio del colegio."""
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["documento", "calificacion"], ["9999999999", 4]]), DOCENTE
        )
        mensaje = resultado["errores"][0]["mensaje"]
        self.assertIn("este curso", mensaje)
        for palabra in ("otro grado", "no está matriculado", "no es una cuenta"):
            self.assertNotIn(palabra, mensaje)

    def test_estudiante_sin_documento_cae_en_el_mensaje_accionable(self):
        """§3.4: Sara existe y está matriculada, pero su ficha no tiene documento."""
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["documento", "calificacion"], ["5555555555", 4]]), DOCENTE
        )
        self.assertIn("usa el correo", resultado["errores"][0]["mensaje"])

    def test_degradacion_una_fila_falla_y_las_demas_se_procesan(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["documento", "calificacion"],
                      ["1023456789", 4.5],
                      ["5555555555", 3.0]]),
            DOCENTE,
        )
        self.assertEqual(len(resultado["filas_validas"]), 1)
        self.assertEqual(len(resultado["errores"]), 1)


class DuplicadosTests(unittest.TestCase):
    """RN-m."""

    def test_dos_filas_del_mismo_estudiante_son_ambas_error(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["correo", "calificacion"],
                      ["ana.gomez@colegio.edu.co", 4.5],
                      ["ana.gomez@colegio.edu.co", 3.0]]),
            DOCENTE,
        )
        # Elegir "la última" en silencio es adivinar cuál nota quería el docente.
        self.assertEqual(resultado["filas_validas"], [])
        self.assertEqual(len(resultado["errores"]), 2)
        self.assertIn("más de una fila", resultado["errores"][0]["mensaje"])

    def test_detecta_el_duplicado_aunque_las_claves_sean_distintas(self):
        # Una fila lo nombra por documento y la otra por correo: sobre el texto
        # del archivo no son iguales, pero son el mismo estudiante.
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["documento", "correo", "calificacion"],
                      ["1023456789", None, 4.5],
                      [None, "ana.gomez@colegio.edu.co", 3.0]]),
            DOCENTE,
        )
        self.assertEqual(resultado["filas_validas"], [])
        self.assertEqual(len(resultado["errores"]), 2)

    def test_un_duplicado_no_arrastra_a_los_demas(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["correo", "calificacion"],
                      ["ana.gomez@colegio.edu.co", 4.5],
                      ["ana.gomez@colegio.edu.co", 3.0],
                      ["luis.pena@colegio.edu.co", 5.0]]),
            DOCENTE,
        )
        self.assertEqual([f["id_estudiante"] for f in resultado["filas_validas"]], [12])


class ResumenTests(unittest.TestCase):
    """RN-l, RN-u, RN-v y la forma de la respuesta."""

    def test_estudiantes_sin_nota_lista_a_quien_el_archivo_no_menciona(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["ana.gomez@colegio.edu.co", 4.5]]), DOCENTE
        )
        sin_nota = {e["id_estudiante"] for e in resultado["estudiantes_sin_nota"]}
        self.assertEqual(sin_nota, {12, 14})

    def test_quien_ya_tenia_nota_no_aparece_como_sin_nota(self):
        # Si a Sara se la pusieron a mano, el aviso sería ruido.
        service = _servicio(_actividad(), notas_existentes=[(14, 3.5)])
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["ana.gomez@colegio.edu.co", 4.5]]), DOCENTE
        )
        sin_nota = {e["id_estudiante"] for e in resultado["estudiantes_sin_nota"]}
        self.assertEqual(sin_nota, {12})

    def test_una_fila_con_error_cuenta_como_mencionado(self):
        # El estudiante está en el archivo aunque su fila haya fallado: no hace
        # falta avisar dos veces de lo mismo.
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["ana.gomez@colegio.edu.co", 99]]), DOCENTE
        )
        sin_nota = {e["id_estudiante"] for e in resultado["estudiantes_sin_nota"]}
        self.assertEqual(sin_nota, {11, 12, 14})

    def test_carga_parcial_convive_con_errores(self):
        """RN-v: lo válido se puede guardar aunque otras filas hayan fallado."""
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["correo", "calificacion"],
                      ["ana.gomez@colegio.edu.co", 4.5],
                      ["luis.pena@colegio.edu.co", 8],
                      ["sara.rios@colegio.edu.co", None]]),
            DOCENTE,
        )
        self.assertEqual(len(resultado["filas_validas"]), 1)
        self.assertEqual(len(resultado["errores"]), 1)
        self.assertEqual(resultado["filas_omitidas"], 1)
        self.assertEqual(resultado["total_filas"], 3)

    def test_los_errores_van_ordenados_por_numero_de_fila(self):
        service = _servicio(_actividad())
        resultado = service.previsualizar_importacion_notas(
            7,
            _archivo([["correo", "calificacion"],
                      ["ajeno@colegio.edu.co", 4.5],
                      ["luis.pena@colegio.edu.co", 99],
                      ["otro.ajeno@colegio.edu.co", 4.5]]),
            DOCENTE,
        )
        self.assertEqual([e["fila"] for e in resultado["errores"]], [2, 3, 4])

    def test_la_respuesta_nombra_la_actividad_destino(self):
        # RN-w: el archivo no lleva id_actividad; el frontend confirma el destino
        # con este nombre dentro del propio botón de guardar.
        service = _servicio(_actividad(nombre="Parcial 1"))
        resultado = service.previsualizar_importacion_notas(
            7, _archivo([["correo", "calificacion"], ["ana.gomez@colegio.edu.co", 4]]), DOCENTE
        )
        self.assertEqual(resultado["actividad"], "Parcial 1")
        self.assertEqual(resultado["id_actividad"], 7)


class SinEfectosTests(unittest.TestCase):
    """RN-q."""

    def test_la_previsualizacion_no_escribe_nada(self):
        service = _servicio(_actividad())
        service.previsualizar_importacion_notas(
            7,
            _archivo([["correo", "calificacion"],
                      ["ana.gomez@colegio.edu.co", 4.5],
                      ["ajeno@colegio.edu.co", 4.5]]),
            DOCENTE,
        )
        service.session.commit.assert_not_called()
        service.session.add.assert_not_called()
        service.session.flush.assert_not_called()
        service.nota_repo.agregar.assert_not_called()


class PlantillaTests(unittest.TestCase):
    """§7.2, §7.3, §8.1."""

    def _plantilla(self, actividad=None, notas_existentes=()):
        service = _servicio(actividad or _actividad(), notas_existentes=notas_existentes)
        contenido, nombre = service.generar_plantilla_notas(7, DOCENTE)
        libro = load_workbook(BytesIO(contenido))
        return libro, nombre

    def test_encabezados_sin_documento(self):
        # Un .xlsx descargado circula fuera de todo control de acceso (RNF-05).
        libro, _ = self._plantilla()
        hoja = libro.worksheets[0]
        encabezados = [celda.value for celda in hoja[1]]
        self.assertEqual(encabezados, ["nombre", "apellido", "correo", "calificacion", "comentario"])
        self.assertNotIn("documento", encabezados)

    def test_ningun_documento_aparece_en_el_archivo(self):
        libro, _ = self._plantilla()
        valores = {
            str(celda.value)
            for fila in libro.worksheets[0].iter_rows()
            for celda in fila
            if celda.value is not None
        }
        for estudiante in ESTUDIANTES:
            if estudiante["documento"]:
                self.assertNotIn(estudiante["documento"], valores)
                self.assertNotIn(estudiante["documento"].replace(".", ""), valores)

    def test_trae_una_fila_por_estudiante_matriculado(self):
        libro, _ = self._plantilla()
        hoja = libro.worksheets[0]
        self.assertEqual(hoja.max_row, 1 + len(ESTUDIANTES))
        correos = {hoja.cell(row=fila, column=3).value for fila in range(2, hoja.max_row + 1)}
        self.assertEqual(correos, {e["correo"] for e in ESTUDIANTES})

    def test_trae_la_calificacion_actual_para_poder_corregir(self):
        libro, _ = self._plantilla(notas_existentes=[(11, 4.5)])
        hoja = libro.worksheets[0]
        notas = {
            hoja.cell(row=fila, column=3).value: hoja.cell(row=fila, column=4).value
            for fila in range(2, hoja.max_row + 1)
        }
        self.assertEqual(notas["ana.gomez@colegio.edu.co"], 4.5)
        self.assertIsNone(notas["sara.rios@colegio.edu.co"])

    def test_encabezados_en_negrita_y_fila_congelada(self):
        libro, _ = self._plantilla()
        hoja = libro.worksheets[0]
        self.assertTrue(hoja["A1"].font.bold)
        self.assertEqual(hoja.freeze_panes, "A2")

    def test_el_correo_va_como_texto(self):
        libro, _ = self._plantilla()
        self.assertEqual(libro.worksheets[0]["C2"].number_format, "@")

    def test_nombre_de_hoja_saneado_y_truncado(self):
        # Excel: máximo 31 caracteres y prohibidos [ ] : * ? / \.
        # openpyxl no siempre avisa y el archivo sale corrupto.
        actividad = _actividad(nombre="Taller: repaso/refuerzo [unidad 3] muy largo de verdad")
        libro, _ = self._plantilla(actividad=actividad)
        titulo = libro.worksheets[0].title
        self.assertLessEqual(len(titulo), 31)
        for prohibido in "[]:*?/\\":
            self.assertNotIn(prohibido, titulo)

    def test_nombre_de_archivo_identifica_la_actividad(self):
        _, nombre = self._plantilla()
        self.assertEqual(nombre, "notas-Matematicas-3-A-Parcial-1.xlsx")
        self.assertTrue(nombre.isascii())

    def test_el_periodo_cerrado_no_bloquea_la_descarga(self):
        # Consultar no es escribir; el bloqueo aplica al importar y al confirmar.
        libro, _ = self._plantilla(actividad=_actividad(estado_periodo="Cerrado"))
        self.assertEqual(libro.worksheets[0].max_row, 1 + len(ESTUDIANTES))

    def test_docente_ajeno_recibe_403(self):
        service = _servicio(_actividad(id_docente=3))
        with self.assertRaises(HTTPException) as exc:
            service.generar_plantilla_notas(7, AJENO)
        self.assertEqual(exc.exception.status_code, 403)

    def test_actividad_inexistente_devuelve_404(self):
        service = _servicio(actividad=None)
        with self.assertRaises(HTTPException) as exc:
            service.generar_plantilla_notas(7, DOCENTE)
        self.assertEqual(exc.exception.status_code, 404)

    def test_la_plantilla_vuelve_a_entrar_por_el_importador(self):
        """El ciclo completo: descargar, escribir las notas y volver a subir."""
        service = _servicio(_actividad())
        contenido, _ = service.generar_plantilla_notas(7, DOCENTE)

        libro = load_workbook(BytesIO(contenido))
        hoja = libro.worksheets[0]
        for fila in range(2, hoja.max_row + 1):
            hoja.cell(row=fila, column=4).value = 4.0
        buffer = BytesIO()
        libro.save(buffer)

        resultado = service.previsualizar_importacion_notas(7, buffer.getvalue(), DOCENTE)
        self.assertEqual(resultado["errores"], [])
        self.assertEqual(len(resultado["filas_validas"]), len(ESTUDIANTES))
        self.assertEqual(resultado["estudiantes_sin_nota"], [])


class RouterTests(unittest.TestCase):
    """Los dos endpoints por HTTP (§8).

    Lo que solo se puede comprobar aquí: que multipart/form-data funciona (o sea,
    que python-multipart está instalado), que la respuesta de la plantilla sale
    como .xlsx con su Content-Disposition, y que la respuesta de importación
    encaja en ImportacionNotasResponse. El servicio se sustituye por un doble:
    su lógica ya está probada arriba.
    """

    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(app)

    def setUp(self):
        self.service = _servicio(_actividad())
        # Las dependencias se guardan como nombres de módulo y no en la clase:
        # asignar una función a un atributo de clase la convierte en método
        # enlazado al leerla con self., y la clave del override deja de ser la
        # función que FastAPI busca.
        app.dependency_overrides[get_current_user] = lambda: DOCENTE
        app.dependency_overrides[get_calificacion_service] = lambda: self.service

    def tearDown(self):
        app.dependency_overrides.clear()

    def test_importar_acepta_multipart_y_devuelve_el_reporte(self):
        contenido = _archivo([["correo", "calificacion"], ["ana.gomez@colegio.edu.co", 4.5]])
        respuesta = self.client.post(
            "/api/notas/importar-excel",
            data={"id_actividad": 7},
            files={"archivo": ("notas.xlsx", contenido, "application/octet-stream")},
        )

        self.assertEqual(respuesta.status_code, 200)
        cuerpo = respuesta.json()
        self.assertEqual(cuerpo["actividad"], "Parcial 1")
        self.assertEqual(cuerpo["filas_validas"][0]["id_estudiante"], 11)
        self.assertEqual(cuerpo["estudiantes_sin_nota"][0]["nombre"], "Luis")

    def test_un_txt_renombrado_a_xlsx_da_400_y_no_500(self):
        respuesta = self.client.post(
            "/api/notas/importar-excel",
            data={"id_actividad": 7},
            files={"archivo": ("notas.xlsx", b"texto plano", "application/octet-stream")},
        )
        self.assertEqual(respuesta.status_code, 400)

    def test_sin_archivo_es_422(self):
        respuesta = self.client.post("/api/notas/importar-excel", data={"id_actividad": 7})
        self.assertEqual(respuesta.status_code, 422)

    def test_plantilla_sale_como_xlsx_descargable(self):
        respuesta = self.client.get("/api/notas/plantilla-excel", params={"id_actividad": 7})

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.headers["content-type"], MEDIA_TYPE_XLSX)
        self.assertIn("attachment", respuesta.headers["content-disposition"])
        self.assertIn("notas-Matematicas-3-A-Parcial-1.xlsx", respuesta.headers["content-disposition"])
        # Y abre de verdad: un .xlsx es un ZIP.
        self.assertTrue(respuesta.content.startswith(b"PK\x03\x04"))

    def test_los_dos_endpoints_requieren_autenticacion(self):
        app.dependency_overrides.clear()
        self.assertIn(
            self.client.get("/api/notas/plantilla-excel", params={"id_actividad": 7}).status_code,
            {401, 403},
        )
        self.assertIn(
            self.client.post(
                "/api/notas/importar-excel",
                data={"id_actividad": 7},
                files={"archivo": ("n.xlsx", b"PK", "application/octet-stream")},
            ).status_code,
            {401, 403},
        )

    def test_el_estudiante_no_puede_importar_ni_descargar(self):
        estudiante = Usuario(id_usuario=11, rol="Estudiante")
        app.dependency_overrides[get_current_user] = lambda: estudiante

        self.assertEqual(
            self.client.get("/api/notas/plantilla-excel", params={"id_actividad": 7}).status_code, 403
        )
        self.assertEqual(
            self.client.post(
                "/api/notas/importar-excel",
                data={"id_actividad": 7},
                files={"archivo": ("n.xlsx", b"PK", "application/octet-stream")},
            ).status_code,
            403,
        )


if __name__ == "__main__":
    unittest.main()
