"""
Pruebas del parser de .xlsx (HU22).

Cubren las reglas que se pueden comprobar sin saber quién es cada estudiante:

  - RN-g: solo .xlsx, validado por la firma de bytes y no por el nombre.
  - RN-h: 2 MB y 1000 filas de datos.
  - RN-i: el archivo necesita al menos una columna de identidad.
  - RN-j: precedencia id_estudiante -> documento -> correo, sin reintentos.
  - RN-l: celda de calificación vacía -> fila omitida, no error.
  - RN-n: se aceptan coma y punto como separador decimal.
  - RN-o: la nota se redondea a 2 decimales.
  - RN-p: comentario de más de 100 caracteres -> error de fila.
  - RN-r: el documento se normaliza (espacios, puntos y guiones).
  - RN-s: notación científica -> error, no se adivina el número.
  - RN-t: el correo se normaliza a minúsculas.

Los archivos se generan en memoria con openpyxl: no hace falta ni Postgres ni
ningún .xlsx guardado en el repositorio.
"""

import unittest
import zipfile
from io import BytesIO

from openpyxl import Workbook

from app.services.importacion_excel import (
    FILAS_MAXIMAS,
    TAMANO_MAXIMO_BYTES,
    ArchivoInvalido,
    parsear_notas_xlsx,
)


def construir_xlsx(filas, titulo="Notas") -> bytes:
    """Un .xlsx real a partir de una lista de listas (la primera son encabezados)."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo
    for fila in filas:
        hoja.append(fila)
    buffer = BytesIO()
    libro.save(buffer)
    libro.close()
    return buffer.getvalue()


class ArchivoTests(unittest.TestCase):
    """RN-g y RN-h."""

    def test_un_txt_renombrado_no_pasa_por_xlsx(self):
        # El nombre y el Content-Type los pone el cliente; la firma de bytes no.
        with self.assertRaises(ArchivoInvalido):
            parsear_notas_xlsx(b"esto no es un excel, es texto plano")

    def test_archivo_vacio(self):
        with self.assertRaises(ArchivoInvalido):
            parsear_notas_xlsx(b"")

    def test_zip_corrupto_no_revienta_con_un_500(self):
        # Empieza con la firma ZIP pero el resto es basura: openpyxl lanza
        # BadZipFile, y tiene que salir como ArchivoInvalido.
        with self.assertRaises(ArchivoInvalido):
            parsear_notas_xlsx(b"PK\x03\x04" + b"\x00" * 200)

    def test_zip_valido_que_no_es_xlsx(self):
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as zf:
            zf.writestr("hola.txt", "contenido")
        with self.assertRaises(ArchivoInvalido):
            parsear_notas_xlsx(buffer.getvalue())

    def test_archivo_mayor_al_limite(self):
        # El tamaño se comprueba sobre los bytes recibidos y antes de abrir el
        # ZIP: un archivo pequeño puede descomprimirse a gigas.
        contenido = b"PK\x03\x04" + b"\x00" * (TAMANO_MAXIMO_BYTES + 1)
        with self.assertRaises(ArchivoInvalido) as exc:
            parsear_notas_xlsx(contenido)
        self.assertIn("MB", str(exc.exception))

    def test_mas_de_mil_filas_de_datos(self):
        filas = [["correo", "calificacion"]]
        filas += [[f"e{i}@colegio.edu.co", 4] for i in range(FILAS_MAXIMAS + 1)]
        with self.assertRaises(ArchivoInvalido) as exc:
            parsear_notas_xlsx(construir_xlsx(filas))
        self.assertIn(str(FILAS_MAXIMAS), str(exc.exception))

    def test_exactamente_mil_filas_pasa(self):
        filas = [["correo", "calificacion"]]
        filas += [[f"e{i}@colegio.edu.co", 4] for i in range(FILAS_MAXIMAS)]
        resultado = parsear_notas_xlsx(construir_xlsx(filas))
        self.assertEqual(len(resultado.filas), FILAS_MAXIMAS)


class EncabezadosTests(unittest.TestCase):
    """RN-i y §7.1."""

    def test_sin_columna_de_identidad_se_rechaza_el_archivo_entero(self):
        # No se empareja por nombre (§2.2): con nombre y apellido no alcanza.
        contenido = construir_xlsx([
            ["nombre", "apellido", "calificacion"],
            ["Ana", "Gómez", 4.5],
        ])
        with self.assertRaises(ArchivoInvalido) as exc:
            parsear_notas_xlsx(contenido)
        self.assertIn("identidad", str(exc.exception))

    def test_sin_columna_de_calificacion_se_rechaza_el_archivo_entero(self):
        contenido = construir_xlsx([["correo"], ["ana@colegio.edu.co"]])
        with self.assertRaises(ArchivoInvalido):
            parsear_notas_xlsx(contenido)

    def test_hoja_vacia(self):
        with self.assertRaises(ArchivoInvalido):
            parsear_notas_xlsx(construir_xlsx([]))

    def test_alias_de_encabezado(self):
        contenido = construir_xlsx([
            ["Cédula", "Nota", "Observación"],
            ["1023456789", "4,5", "Bien"],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.errores, [])
        self.assertEqual(resultado.filas[0].clave, "documento")
        self.assertEqual(resultado.filas[0].comentario, "Bien")

    def test_encabezados_con_tildes_espacios_y_mayusculas(self):
        contenido = construir_xlsx([
            ["  Correo Electrónico ", "CALIFICACIÓN"],
            ["ana@colegio.edu.co", 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(len(resultado.filas), 1)
        self.assertEqual(resultado.filas[0].clave, "correo")

    def test_columnas_desconocidas_se_ignoran_sin_error(self):
        # Los docentes llevan columnas propias en sus planillas.
        contenido = construir_xlsx([
            ["correo", "puesto", "calificacion", "acudiente"],
            ["ana@colegio.edu.co", 3, 4.5, "María"],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.errores, [])
        self.assertEqual(resultado.filas[0].calificacion, 4.5)


class IdentidadTests(unittest.TestCase):
    """RN-j, RN-r, RN-s, RN-t."""

    def test_precedencia_id_documento_correo(self):
        contenido = construir_xlsx([
            ["id_estudiante", "documento", "correo", "calificacion"],
            [12, "1023456789", "ana@colegio.edu.co", 4],
            [None, "1023456789", "ana@colegio.edu.co", 4],
            [None, None, "ana@colegio.edu.co", 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual([f.clave for f in resultado.filas], ["id_estudiante", "documento", "correo"])
        self.assertEqual(resultado.filas[0].valor, 12)

    def test_id_estudiante_no_numerico_no_reintenta_con_el_correo(self):
        # RN-j: probar todas las claves hasta que alguna pegue es cómo se le
        # termina poniendo la nota de un estudiante a otro.
        contenido = construir_xlsx([
            ["id_estudiante", "correo", "calificacion"],
            ["ABC", "ana@colegio.edu.co", 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.filas, [])
        self.assertEqual(resultado.errores[0].columna, "id_estudiante")

    def test_documento_se_normaliza(self):
        contenido = construir_xlsx([
            ["documento", "calificacion"],
            ["1.023.456.789", 4],
            ["  1023-456-789  ", 4],
            ["0012345678", 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual([f.valor for f in resultado.filas],
                         ["1023456789", "1023456789", "0012345678"])

    def test_documento_numerico_no_llega_con_punto_cero(self):
        # Excel entrega las celdas numéricas como float: sin convertirlo a int
        # se buscaría el documento "1023456789.0".
        contenido = construir_xlsx([["documento", "calificacion"], [1023456789, 4]])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.filas[0].valor, "1023456789")

    def test_notacion_cientifica_es_error_y_no_se_adivina(self):
        contenido = construir_xlsx([
            ["documento", "calificacion"],
            ["1,02E+09", 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.filas, [])
        self.assertIn("notación científica", resultado.errores[0].mensaje)

    def test_correo_se_normaliza_a_minusculas(self):
        contenido = construir_xlsx([
            ["correo", "calificacion"],
            ["  Ana.Gomez@Colegio.Edu.CO ", 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.filas[0].valor, "ana.gomez@colegio.edu.co")

    def test_fila_sin_ninguna_clave_es_error_de_fila(self):
        contenido = construir_xlsx([
            ["nombre", "correo", "calificacion"],
            ["Ana", None, 4],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.errores[0].columna, "identidad")
        self.assertEqual(resultado.errores[0].fila, 2)


class CalificacionTests(unittest.TestCase):
    """RN-l, RN-n, RN-o, RN-a."""

    def _una_fila(self, calificacion):
        return parsear_notas_xlsx(construir_xlsx([
            ["correo", "calificacion"],
            ["ana@colegio.edu.co", calificacion],
        ]))

    def test_celda_vacia_se_omite_en_silencio(self):
        resultado = self._una_fila(None)
        self.assertEqual(resultado.filas, [])
        self.assertEqual(resultado.errores, [])
        self.assertEqual(resultado.omitidas, 1)

    def test_coma_y_punto_son_equivalentes(self):
        self.assertEqual(self._una_fila("3,2").filas[0].calificacion, 3.2)
        self.assertEqual(self._una_fila("3.2").filas[0].calificacion, 3.2)
        self.assertEqual(self._una_fila(3.2).filas[0].calificacion, 3.2)

    def test_se_redondea_a_dos_decimales(self):
        # La columna es NUMERIC(3,2). Se redondea aquí para que lo que el
        # docente aprueba en la vista previa sea lo que se guarda.
        self.assertEqual(self._una_fila("4.567").filas[0].calificacion, 4.57)

    def test_fuera_de_rango_es_error_de_fila(self):
        for valor in (8, -1, 5.01):
            resultado = self._una_fila(valor)
            self.assertEqual(resultado.filas, [])
            self.assertEqual(resultado.errores[0].columna, "calificacion")

    def test_texto_no_numerico_es_error_de_fila(self):
        resultado = self._una_fila("excelente")
        self.assertEqual(resultado.errores[0].columna, "calificacion")

    def test_booleano_no_cuenta_como_uno(self):
        # bool es subclase de int en Python: sin la guarda, TRUE valdría 1.00.
        resultado = self._una_fila(True)
        self.assertEqual(resultado.filas, [])
        self.assertEqual(resultado.errores[0].columna, "calificacion")

    def test_fila_omitida_no_valida_la_identidad(self):
        # Subir la plantilla a medio llenar tiene que funcionar aunque las filas
        # en blanco traigan un correo que no se va a usar.
        contenido = construir_xlsx([
            ["correo", "calificacion"],
            ["ana@colegio.edu.co", 4],
            ["no-es-un-correo", None],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(len(resultado.filas), 1)
        self.assertEqual(resultado.errores, [])
        self.assertEqual(resultado.omitidas, 1)


class ComentarioTests(unittest.TestCase):
    """RN-p."""

    def test_comentario_largo_es_error_y_no_se_trunca(self):
        contenido = construir_xlsx([
            ["correo", "calificacion", "comentario"],
            ["ana@colegio.edu.co", 4, "x" * 101],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.filas, [])
        self.assertEqual(resultado.errores[0].columna, "comentario")

    def test_comentario_de_cien_caracteres_pasa(self):
        contenido = construir_xlsx([
            ["correo", "calificacion", "comentario"],
            ["ana@colegio.edu.co", 4, "x" * 100],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(len(resultado.filas[0].comentario), 100)


class FormaDelResultadoTests(unittest.TestCase):

    def test_numeros_de_fila_son_los_de_excel(self):
        # Un mensaje que dice "fila 5" cuando en Excel es la 6 es peor que no
        # decir nada: el encabezado es la fila 1 y el primer dato la 2.
        contenido = construir_xlsx([
            ["correo", "calificacion"],
            ["ana@colegio.edu.co", 4],
            ["luis@colegio.edu.co", 99],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.filas[0].fila, 2)
        self.assertEqual(resultado.errores[0].fila, 3)

    def test_filas_en_blanco_del_final_no_cuentan(self):
        contenido = construir_xlsx([
            ["correo", "calificacion"],
            ["ana@colegio.edu.co", 4],
            [None, None],
            [None, None],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual(resultado.total_filas, 1)
        self.assertEqual(resultado.omitidas, 0)

    def test_total_filas_suma_validas_errores_y_omitidas(self):
        contenido = construir_xlsx([
            ["correo", "calificacion"],
            ["ana@colegio.edu.co", 4],
            ["luis@colegio.edu.co", 99],
            ["sara@colegio.edu.co", None],
        ])
        resultado = parsear_notas_xlsx(contenido)
        self.assertEqual((len(resultado.filas), len(resultado.errores), resultado.omitidas), (1, 1, 1))
        self.assertEqual(resultado.total_filas, 3)

    def test_solo_se_lee_la_primera_hoja(self):
        libro = Workbook()
        libro.active.title = "Parcial 1"
        libro.active.append(["correo", "calificacion"])
        libro.active.append(["ana@colegio.edu.co", 4])
        segunda = libro.create_sheet("Otra")
        segunda.append(["correo", "calificacion"])
        segunda.append(["luis@colegio.edu.co", 5])
        buffer = BytesIO()
        libro.save(buffer)
        libro.close()

        resultado = parsear_notas_xlsx(buffer.getvalue())
        self.assertEqual(len(resultado.filas), 1)
        self.assertEqual(resultado.filas[0].valor, "ana@colegio.edu.co")

    def test_el_parser_no_conoce_estudiantes(self):
        """El parser traduce, no resuelve: devuelve la clave, no un id."""
        contenido = construir_xlsx([
            ["documento", "calificacion"],
            ["1023456789", 4],
        ])
        fila = parsear_notas_xlsx(contenido).filas[0]
        self.assertFalse(hasattr(fila, "id_estudiante"))


if __name__ == "__main__":
    unittest.main()
